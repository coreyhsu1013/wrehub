from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from datetime import date

from django.core.management.base import BaseCommand, CommandError

from hub.models import UrbanRenewalCase, UrbanRenewalBonus

PING_SQM = Decimal("3.305785")

# 107/7/6  or 107/07/06
RE_ROC_DATE = re.compile(r"^(?P<y>\d{2,3})/(?P<m>\d{1,2})/(?P<d>\d{1,2})$")

def parse_roc_date(s: str) -> date | None:
    """解析民國日期格式：107/7/6 或 107/07/06"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    m = RE_ROC_DATE.match(s)
    if not m:
        return None
    y = int(m.group("y")) + 1911
    mo = int(m.group("m"))
    d = int(m.group("d"))
    try:
        return date(y, mo, d)
    except Exception:
        return None

def to_decimal(s) -> Decimal | None:
    """轉換為 Decimal，處理 None、空字串、NaN"""
    if s is None:
        return None
    # 處理 pandas NaN
    try:
        import math
        if isinstance(s, float) and math.isnan(s):
            return None
    except ImportError:
        pass
    if isinstance(s, str):
        s = s.strip()
        if not s or s.lower() == "nan":
            return None
        s = s.replace(",", "").replace("㎡", "").replace("％", "").replace("%", "")
    try:
        return Decimal(str(s))
    except Exception:
        return None

def parse_parcel(parcel_text: str) -> tuple[str, str, str]:
    """
    從 parcel 原文字串解析 section, subsection, parcel_no
    例如：「大安段一小段292」-> ("大安段", "一小段", "292")
    """
    if not parcel_text:
        return ("", "", "")
    
    parcel_text = str(parcel_text).strip()
    section = ""
    subsection = ""
    parcel_no = ""
    
    # 找 section：以「XX段」抓取
    section_match = re.search(r"(.+?段)", parcel_text)
    if section_match:
        section = section_match.group(1)
    
    # 找 subsection：以「X小段」抓取（例如 四小段、一小段）
    subsection_match = re.search(r"([一二三四五六七八九十\d]+小段)", parcel_text)
    if subsection_match:
        subsection = subsection_match.group(1)
    
    # 找 parcel_no：抓「數字地號」(例如 292)
    parcel_match = re.search(r"(\d+)", parcel_text)
    if parcel_match:
        parcel_no = parcel_match.group(1)
    
    return (section, subsection, parcel_no)

# Bonus 欄位映射
BONUS_FIELD_MAP = {
    "bonus_structure_eval_pct": "structure",
    "bonus_seismic_pct": "seismic",
    "bonus_green_pct": "green",
    "bonus_smart_pct": "smart",
    "bonus_accessible_pct": "barrierfree",
    "bonus_public_facility_donation_pct": "donation",
    "bonus_schedule_scale_pct": "schedule",
    "bonus_setback_pct": "other",
}

class Command(BaseCommand):
    help = "Import Taipei urban renewal 1049 cases from Excel file."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to xlsx file")
        parser.add_argument("--clear", action="store_true", help="Clear existing cases (source='taipei_ur')")
        parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
        parser.add_argument("--limit", type=int, default=0, help="Only process first N rows (0=all)")

    def handle(self, *args, **opts):
        file_path = Path(opts["file"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        dry_run = bool(opts["dry_run"])
        limit = int(opts["limit"] or 0)
        clear = bool(opts["clear"])

        # 嘗試導入 pandas，如果沒有則用 openpyxl
        try:
            import pandas as pd
            use_pandas = True
        except ImportError:
            try:
                from openpyxl import load_workbook
                use_pandas = False
            except ImportError:
                raise CommandError("Need either pandas or openpyxl installed. Try: pip install pandas")

        # 清空現有資料
        if clear and not dry_run:
            deleted_count = UrbanRenewalCase.objects.filter(source="taipei_ur").count()
            UrbanRenewalCase.objects.filter(source="taipei_ur").delete()
            self.stdout.write(self.style.WARNING(f"Cleared {deleted_count} existing cases"))

        stats = {
            "processed": 0,
            "created": 0,
            "updated": 0,
            "bonus_created": 0,
            "bonus_updated": 0,
            "skipped": 0,
        }

        # 讀取 Excel
        if use_pandas:
            try:
                df = pd.read_excel(file_path)
                rows = df.to_dict("records")
            except Exception as e:
                raise CommandError(f"Failed to read Excel with pandas: {e}")
        else:
            # 使用 openpyxl
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
            except Exception as e:
                raise CommandError(f"Failed to read Excel with openpyxl: {e}")

        # 處理每一列
        for idx, row in enumerate(rows):
            if limit and idx >= limit:
                break

            # 取得必要欄位
            case_seq = row.get("case_seq")
            if case_seq is None:
                stats["skipped"] += 1
                continue

            # 轉換 case_seq 為整數
            try:
                case_seq = int(case_seq)
            except (ValueError, TypeError):
                stats["skipped"] += 1
                continue

            stats["processed"] += 1

            # 解析欄位
            district = str(row.get("district", "")).strip()
            address = str(row.get("address", "")).strip()
            parcel_text = str(row.get("parcel", "")).strip()
            note = str(row.get("note", "")).strip()
            
            # 解析日期
            approved_date = parse_roc_date(str(row.get("approved_date", "")))
            
            # 解析面積
            site_area_sqm = to_decimal(row.get("site_area_sqm"))
            site_area_ping = None
            if site_area_sqm is not None:
                try:
                    site_area_ping = (site_area_sqm / PING_SQM).quantize(Decimal("0.01"))
                except Exception:
                    pass
            
            # 解析總獎勵
            total_bonus_pct = to_decimal(row.get("total_bonus_pct_num"))
            
            # 解析地段地號
            section, subsection, parcel_no = parse_parcel(parcel_text)
            
            # 準備 raw 資料（整列轉 dict）
            raw_data = {}
            for key, value in row.items():
                if value is None:
                    continue
                # 處理 pandas NaN
                try:
                    import math
                    if isinstance(value, float) and math.isnan(value):
                        continue
                except ImportError:
                    pass
                # 處理空字串
                if isinstance(value, str) and not value.strip():
                    continue
                raw_data[str(key)] = value
            
            # 準備 defaults
            defaults = {
                "source": "taipei_ur",
                "district": district,
                "address": address,
                "approved_date": approved_date,
                "site_area_sqm": site_area_sqm,
                "site_area_ping": site_area_ping,
                "total_bonus_pct": total_bonus_pct,
                "section": section,
                "subsection": subsection,
                "parcel_no": parcel_no,
                "raw": raw_data,
            }

            if dry_run:
                # 只計數，不寫入
                stats["created"] += 1
                # 計算 bonus 數量
                for field_name, code in BONUS_FIELD_MAP.items():
                    bonus_value = to_decimal(row.get(field_name))
                    if bonus_value is not None:
                        stats["bonus_created"] += 1
                continue

            # 寫入 UrbanRenewalCase
            case, is_created = UrbanRenewalCase.objects.update_or_create(
                source="taipei_ur",
                case_no=case_seq,
                defaults=defaults,
            )
            
            if is_created:
                stats["created"] += 1
            else:
                stats["updated"] += 1

            # 寫入 UrbanRenewalBonus
            for field_name, code in BONUS_FIELD_MAP.items():
                bonus_value = to_decimal(row.get(field_name))
                if bonus_value is not None:
                    bonus, bonus_created = UrbanRenewalBonus.objects.update_or_create(
                        case=case,
                        code=code,
                        defaults={"bonus_pct": bonus_value},
                    )
                    if bonus_created:
                        stats["bonus_created"] += 1
                    else:
                        stats["bonus_updated"] += 1

        # 輸出統計
        self.stdout.write(self.style.SUCCESS(
            f"Done. processed={stats['processed']} "
            f"created={stats['created']} updated={stats['updated']} "
            f"bonus_created={stats['bonus_created']} bonus_updated={stats['bonus_updated']} "
            f"skipped={stats['skipped']} dry_run={dry_run}"
        ))
